from __future__ import annotations

import csv
import json
import os
import time
from collections.abc import Sequence
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from common_config.config.settings import get_settings
from common_config.db.connection import MongoDBConnection  # type: ignore
from common_config.utils.logger import get_logger

logger = get_logger(__name__)


# ----------------------- Data Classes -----------------------
@dataclass(slots=True)
class ValidationSummary:
    total_rows: int = 0
    matched: int = 0
    not_found: int = 0
    mismatched: int = 0

    def to_dict(self):
        return asdict(self)


@dataclass(slots=True)
class MismatchRecord:
    """Represents a mismatched record with database values and mismatched fields."""

    HcmId: str
    FirstName: str
    LastName: str
    Dob: str
    Gender: str
    MismatchedField: str  # Comma-separated field names


MANDATORY_FIELDS = ["HcmId", "FirstName", "LastName", "Dob", "Gender"]
COMPARISON_FIELDS = ["FirstName", "LastName", "Dob", "Gender"]  # exclude HcmId because it is key only


# ----------------------- Excel Support -----------------------
class ExcelDictReader:
    """Excel reader that mimics csv.DictReader interface."""

    def __init__(self, file_path: Path, sheet_name: str | None = None):
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not available for Excel support")

        self.workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        self.sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        self.headers: list[str] = []
        self._rows_iter = None
        self._current_row = 1

        # Read headers from first row
        if self.sheet is not None:
            first_row = next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            self.headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(first_row)]
            self._rows_iter = self.sheet.iter_rows(min_row=2, values_only=True)

    def __iter__(self):
        return self

    def __next__(self):
        if self._rows_iter is None:
            raise StopIteration
        try:
            row_values = next(self._rows_iter)
            self._current_row += 1

            # Convert row to dictionary, handling None values and date formatting
            row_dict = {}
            for i, value in enumerate(row_values):
                if i < len(self.headers):
                    key = self.headers[i]
                    if value is None:
                        row_dict[key] = ""
                    elif isinstance(value, datetime) and key == "Dob":
                        # Excel dates are already parsed - format as YYYY-MM-DD
                        row_dict[key] = value.strftime("%Y-%m-%d")
                    else:
                        row_dict[key] = str(value)

            return row_dict

        except StopIteration:
            self.workbook.close()
            raise

    @property
    def fieldnames(self):
        return self.headers


# ----------------------- Normalization Helpers -----------------------
def normalize_name(value: str | None) -> str:
    return (value or "").strip().upper()


def normalize_dob_excel(value: str | None) -> str:
    """Simplified DOB normalization for Excel - expects YYYY-MM-DD format."""
    if not value:
        return ""
    raw = value.strip()
    if not raw:
        return ""

    # For Excel, DOB should already be formatted as YYYY-MM-DD
    # Just validate the format and return as-is
    if len(raw) >= 10 and raw[4] == "-" and raw[7] == "-":
        try:
            # Validate it's a proper date
            datetime.strptime(raw[:10], "%Y-%m-%d")
            return raw[:10]
        except ValueError:
            # If validation fails, return as-is (will be flagged as mismatch)
            return raw[:10]

    # If not in expected format, return as-is
    return raw[:10] if len(raw) >= 10 else raw


def normalize_gender(value: str | None) -> str:
    return (value or "").strip()


def _build_normalizers_excel():
    """Build normalizers for Excel processing - simplified DOB handling."""
    return {
        "FirstName": normalize_name,
        "LastName": normalize_name,
        "Dob": normalize_dob_excel,
        "Gender": normalize_gender,
    }


def _resolve_mongo(env_suffix: str) -> tuple[str, str]:
    """Resolve Mongo URI and database using shared settings with APP_ENV.

    Sets APP_ENV then loads settings (which merges shared and local .env files and
    resolves suffixed variables). Returns (uri, database).
    """
    os.environ["APP_ENV"] = env_suffix.upper()
    settings = get_settings()
    if not settings.mongodb_uri or not settings.database_name:
        raise RuntimeError(
            "Mongo configuration unresolved: ensure MONGODB_URI_<ENV> and DATABASE_NAME_<ENV> or base vars are set"
        )
    return settings.mongodb_uri, settings.database_name


# ----------------------- Core Logic -----------------------
def _extract_db_field(doc: dict, field: str) -> Any:
    value = doc.get(field)
    if field == "Dob" and value is not None:
        # value may be datetime; normalize to YYYY-MM-DD
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        # If stored as string ensure truncated
        if isinstance(value, str):
            return value[:10]
    return value


def _compare_row(row: dict, doc: dict, normalizers: dict[str, Any]) -> list[str]:
    mismatches: list[str] = []
    for field in COMPARISON_FIELDS:
        csv_val_raw = row.get(field, "")
        if not csv_val_raw:
            mismatches.append(f"{field} missing on csv")
            continue
        norm_fn = normalizers[field]
        csv_val = norm_fn(csv_val_raw)
        db_val_raw = _extract_db_field(doc, field)
        db_val = norm_fn(str(db_val_raw)) if db_val_raw is not None else ""
        if csv_val != db_val:
            mismatches.append(field)
    return mismatches


def _canonicalize_headers(raw_fields: Sequence[str] | None, debug: bool) -> tuple[list[str], dict[str, str], bool]:
    """Return (canonical_fields, raw_to_canonical_map, legacy_comments_present).

    Canonicalization rules:
      * Strip BOM (\ufeff), leading/trailing whitespace
      * Case-insensitive match for mandatory fields (HcmId, FirstName, LastName, Dob, Gender)
      * Map 'Mismatched Fields' -> 'Comments'
    """
    legacy = False
    if not raw_fields:
        return [], {}, legacy
    mapping: dict[str, str] = {}
    canonical: list[str] = []
    mandatory_lower = {
        "hcmid": "HcmId",
        "firstname": "FirstName",
        "lastname": "LastName",
        "dob": "Dob",
        "gender": "Gender",
    }
    for raw in raw_fields:
        cleaned = raw.lstrip("\ufeff").strip()
        low = cleaned.lower()
        if low == "mismatched fields":
            canon = "Comments"
            legacy = True
        elif low in mandatory_lower:
            canon = mandatory_lower[low]
        elif cleaned in ("Match Found", "Comments"):
            canon = cleaned
        else:
            canon = cleaned  # passthrough for any extra columns
        mapping[raw] = canon
        canonical.append(canon)
    # Ensure required output columns present
    if "Match Found" not in canonical:
        canonical.append("Match Found")
    if "Comments" not in canonical:
        canonical.append("Comments")
    if debug:
        logger.debug("Header canonicalization raw->canon: %s", mapping)
    return canonical, mapping, legacy


def _write_mismatches_csv(output_csv: Path, mismatches: list[MismatchRecord], debug: bool) -> Path:
    """Write mismatches to a separate CSV file with database values.

    Args:
        output_csv: The main output CSV path (used to derive mismatches filename)
        mismatches: List of mismatch records with database values
        debug: Enable debug logging

    Returns:
        Path to the mismatches CSV file
    """
    # Derive mismatches filename from output filename
    output_stem = output_csv.stem  # e.g., "20251013_190355_output"
    if output_stem.endswith("_output"):
        mismatches_stem = output_stem + "_mismatches"
    else:
        mismatches_stem = output_stem + "_mismatches"

    mismatches_csv = output_csv.parent / f"{mismatches_stem}.csv"

    if debug:
        logger.debug("Writing %d mismatch records to %s", len(mismatches), mismatches_csv)

    with mismatches_csv.open("w", newline="", encoding="utf-8") as f:
        fieldnames = ["HcmId", "FirstName", "LastName", "Dob", "Gender", "Mismatched Field"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for mismatch in mismatches:
            writer.writerow(
                {
                    "HcmId": mismatch.HcmId,
                    "FirstName": mismatch.FirstName,
                    "LastName": mismatch.LastName,
                    "Dob": mismatch.Dob,
                    "Gender": mismatch.Gender,
                    "Mismatched Field": mismatch.MismatchedField,
                }
            )

    logger.info("Mismatches CSV written: %s (%d records)", mismatches_csv, len(mismatches))
    return mismatches_csv


def run_validation_excel(
    input_file: Path,
    output_csv: Path,
    collection_name: str,
    batch_size: int | None,
    limit: int | None,
    dry_run: bool,
    progress_every: int,
    mongodb_env: str,
    max_retries: int = 3,
    base_retry_sleep: float = 0.5,
    excel_sheet: str | None = None,
    debug: bool = False,
) -> dict:
    """Excel validation with batch Mongo lookups using $in queries.

    Steps:
      1. Read Excel header & prepare CSV writer.
      2. Stream rows, accumulating batches by HcmId.
      3. For each batch, perform one query: { HcmId: { $in: [...] } }.
      4. Compare & write results to CSV.
      5. Write separate mismatches CSV with database values.
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl not available - cannot process Excel files")

    logger.info("Starting Excel validation - Input file: %s", input_file)
    logger.info("Output file: %s, Collection: %s, Environment: %s", output_csv, collection_name, mongodb_env)
    if excel_sheet:
        logger.info("Excel sheet: %s", excel_sheet)

    summary = ValidationSummary()
    normalizers = _build_normalizers_excel()
    mismatches: list[MismatchRecord] = []  # Collect all mismatched records

    # Create Excel reader
    excel_reader = ExcelDictReader(input_file, excel_sheet)
    raw_header = excel_reader.fieldnames

    canonical_fields, header_map, legacy_comments_present = _canonicalize_headers(raw_header, debug)
    out_fields = canonical_fields

    if dry_run:
        logger.info("Dry run: output file will not be written")
        f_out = None
    else:
        output_csv.parent.mkdir(parents=True, exist_ok=True)
        f_out = output_csv.open("w", newline="", encoding="utf-8")
    writer = csv.DictWriter(f_out, fieldnames=out_fields) if f_out else None
    if writer:
        writer.writeheader()

    # Derive effective batch size
    if batch_size:
        effective_batch = batch_size
    else:
        try:
            s = get_settings()
            proc = getattr(s, "processing", None)
            effective_batch = getattr(proc, "batch_size", 1000) if proc else 1000
        except Exception:
            effective_batch = 1000

    mongo_uri, mongo_db = _resolve_mongo(mongodb_env)
    connection = MongoDBConnection(
        type(
            "Cfg",
            (),
            {
                "get_mongodb_uri": lambda self=mongo_uri: mongo_uri,  # type: ignore
                "get_mongodb_database": lambda self=mongo_db: mongo_db,  # type: ignore
            },
        )()
    )
    with connection:
        db = connection.get_database()
        collection = db[collection_name]

        # Streaming read
        batch_rows: list[dict[str, str]] = []
        for i, row in enumerate(excel_reader, start=1):
            # Remap row keys to canonical
            canonical_row: dict[str, str] = {}
            for raw_key, value in row.items():
                canon_key = header_map.get(raw_key, raw_key)
                if canon_key is not None:  # Type guard: ensure canon_key is not None
                    canonical_row[canon_key] = value if value is not None else ""
            # Map legacy header per row if needed
            if legacy_comments_present and "Comments" not in canonical_row and "Mismatched Fields" in canonical_row:
                canonical_row["Comments"] = canonical_row.get("Mismatched Fields", "")
                canonical_row.pop("Mismatched Fields", None)
            if limit and i > limit:
                break
            batch_rows.append(canonical_row)
            if len(batch_rows) >= effective_batch:
                _process_batch(
                    batch_rows,
                    collection,
                    writer,
                    summary,
                    max_retries,
                    base_retry_sleep,
                    debug,
                    normalizers,
                    mismatches,
                )
                batch_rows = []
            if i % progress_every == 0:
                logger.info(
                    "Processed %d rows (matched=%d mismatched=%d not_found=%d)",
                    i,
                    summary.matched,
                    summary.mismatched,
                    summary.not_found,
                )
        # leftover
        if batch_rows:
            _process_batch(
                batch_rows,
                collection,
                writer,
                summary,
                max_retries,
                base_retry_sleep,
                debug,
                normalizers,
                mismatches,
            )

    if f_out:
        f_out.close()

    # Write mismatches CSV if not dry run
    mismatches_csv_path = None
    if not dry_run and mismatches:
        mismatches_csv_path = _write_mismatches_csv(output_csv, mismatches, debug)

    final_summary = summary.to_dict()
    if mismatches_csv_path:
        final_summary["mismatches_csv"] = str(mismatches_csv_path)
    logger.info("Summary: %s", final_summary)
    # Also emit JSON on stdout for machine consumption when not dry-run
    try:
        print(json.dumps({"summary": final_summary}, separators=(",", ":")))
    except Exception:
        pass
    return final_summary


def _process_batch(
    rows: list[dict[str, str]],
    collection,  # pymongo collection
    writer: csv.DictWriter | None,
    summary: ValidationSummary,
    max_retries: int,
    base_retry_sleep: float,
    debug: bool,
    normalizers: dict[str, Any],
    mismatches: list[MismatchRecord],
) -> None:
    # Collect HcmIds for rows that have all mandatory fields present
    hcmids: list[str] = []
    for r in rows:
        missing_fields = [f for f in MANDATORY_FIELDS if not r.get(f)]
        if missing_fields:
            # Mark immediately
            r["Match Found"] = "False"
            r["Comments"] = ",".join(f"{mf} missing on csv" for mf in missing_fields)
            summary.mismatched += 1
            summary.total_rows += 1
            if debug:
                logger.debug(
                    "Row flagged missing mandatory fields=%s raw_keys=%s sample_values=%s",
                    missing_fields,
                    list(r.keys()),
                    {k: repr(r.get(k)) for k in MANDATORY_FIELDS},
                )
        else:
            hcmids.append(r["HcmId"].strip())
    # Query only for rows not already classified as missing mandatory
    if hcmids:
        attempt = 0
        if debug:
            logger.debug("Issuing batch find on HcmId list (count=%d)", len(hcmids))
        while True:
            try:
                query = {"HcmId": {"$in": hcmids}}
                projection = dict.fromkeys(MANDATORY_FIELDS, 1)
                if debug:
                    logger.debug("Mongo query=%s projection=%s first_ids=%s", query, projection, hcmids[:5])
                docs = list(collection.find(query, projection))
                break
            except Exception as e:  # pragma: no cover - external dependency
                attempt += 1
                if attempt > max_retries:
                    logger.error("Batch query failed after %d attempts: %s", attempt - 1, e)
                    docs = []
                    break
                sleep_for = base_retry_sleep * (2 ** (attempt - 1))
                logger.warning(
                    "Batch query failed (attempt %d/%d): %s; retrying in %.2fs", attempt, max_retries, e, sleep_for
                )
                time.sleep(sleep_for)
        doc_map = {d.get("HcmId"): d for d in docs}
    else:
        doc_map = {}

    for r in rows:
        if "Match Found" in r and r["Match Found"] == "False" and r.get("Comments"):
            # Already processed due to missing fields
            if writer:
                writer.writerow(r)
            continue
        summary.total_rows += 1
        if "Match Found" not in r:
            r["Match Found"] = "False"
        if "Comments" not in r:
            r["Comments"] = ""
        hcmid = r.get("HcmId", "").strip()
        if not hcmid:
            r["Comments"] = "HcmId missing on db"
            summary.mismatched += 1
            if writer:
                writer.writerow(r)
            continue
        doc = doc_map.get(hcmid)
        if not doc:
            r["Match Found"] = "False"
            r["Comments"] = "HcmId Not Found"
            summary.not_found += 1
            # For HcmId Not Found: use input data values with "HcmId" as mismatched field
            mismatches.append(
                MismatchRecord(
                    HcmId=r.get("HcmId", ""),
                    FirstName=r.get("FirstName", ""),
                    LastName=r.get("LastName", ""),
                    Dob=r.get("Dob", ""),
                    Gender=r.get("Gender", ""),
                    MismatchedField="HcmId",
                )
            )
            if debug:
                logger.debug("Unmatched HcmId=%s (not found in batch map)", hcmid)
        else:
            mismatches_list = _compare_row(r, doc, normalizers)
            if not mismatches_list:
                r["Match Found"] = "True"
                r["Comments"] = ""
                summary.matched += 1
            else:
                r["Match Found"] = "False"
                r["Comments"] = ",".join(mismatches_list)
                summary.mismatched += 1
                # For field mismatches: use database values with comma-separated mismatched fields
                mismatches.append(
                    MismatchRecord(
                        HcmId=str(_extract_db_field(doc, "HcmId") or ""),
                        FirstName=str(_extract_db_field(doc, "FirstName") or ""),
                        LastName=str(_extract_db_field(doc, "LastName") or ""),
                        Dob=str(_extract_db_field(doc, "Dob") or ""),
                        Gender=str(_extract_db_field(doc, "Gender") or ""),
                        MismatchedField=",".join(mismatches_list),
                    )
                )
                if debug:
                    logger.debug(
                        "Mismatched HcmId=%s mismatches=%s row_values=%s db_values=%s",
                        hcmid,
                        mismatches_list,
                        {f: r.get(f) for f in MANDATORY_FIELDS},
                        {f: _extract_db_field(doc, f) for f in MANDATORY_FIELDS},
                    )
        if writer:
            writer.writerow(r)
